"""
PriceLens — Google Sheets Edition  (v4)
═══════════════════════════════════════════════════════════════
Sources: JioMart · BigBasket · DMart Ready
Master Mapper Sheet: BenchMark | Superkid | Comp_id | url | pincode
RESULT cols: run_id | run_at | product_name | is_critical
             comp_mrp | comp_sp | comp_pct_off
             sk_mrp | sk_sp | sk_pct_off
             diff_pct_off | diff_sp | mode | status

New in v4:
  • DMart Ready scraper (parses Next.js __NEXT_F payload)
  • Source-level run filter: run only jio / bb / dmart / all
  • Pause / Resume / Stop mid-run
  • Selective Superkid run
  • Lookup endpoint
  • Inline results back into Master Mapper

v4.1 fix:
  • JioMart now uses curl_cffi (chrome110 impersonation) + cookies
    matching the working standalone script exactly.
  • Fake x-fp-signature removed; real authorization token used as-is.
  • jio_cookies_json config key added (same pattern as BB/DMart).
"""
from dotenv import load_dotenv
load_dotenv()
import os, threading, traceback, time, re, json, io
from datetime import datetime
from collections import deque
from flask import Flask, render_template, request, jsonify, send_file

try:
    from curl_cffi import requests as curl_requests
    CURL_AVAILABLE = True
except ImportError:
    import requests as curl_requests
    CURL_AVAILABLE = False

try:
    import requests as req_lib
except ImportError:
    req_lib = None

try:
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    GOOGLE_AVAILABLE = True
except ImportError:
    GOOGLE_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

run_log = deque(maxlen=3000)

_pause_event = threading.Event()
_stop_event  = threading.Event()
_pause_event.set()
_stop_event.clear()

active_run = {
    "running": False, "paused": False, "stopped": False,
    "progress": 0, "total": 0, "current": "",
    "run_id": "", "mode": "", "superkids": [],
    "source_filter": "all",
}
run_history = []

SCOPES     = ["https://www.googleapis.com/auth/spreadsheets"]
TOKEN_PATH = "token.json"

MASTER_MAPPER_ID    = os.getenv("MASTER_MAPPER_ID", "1KmTy3xrazd0a8Ew1HUOVvnxZv0QaT6ARR4aiDYZuQVM")
MASTER_MAPPER_SHEET = os.getenv("MASTER_MAPPER_SHEET", "Master mapper")


INPUT_COLS = ["BenchMark", "Superkid", "Comp_id", "url", "pincode"]

RESULT_COLS = [
    "run_id", "run_at", "product_name", "is_critical", "mode",
    "comp_mrp", "comp_sp", "comp_pct_off",
    "sk_mrp",  "sk_sp",  "sk_pct_off",
    "diff_pct_off", "diff_sp", "status",
]

JIOMART_SEARCH_URL = "https://www.jiomart.com/ext/vertex/application/api/v1.0/products"

JIO_BASE_HEADERS = {
    "accept":               "application/json, text/plain, */*",
    "accept-language":      "en-GB,en-US;q=0.9,en;q=0.8",
    "sec-ch-ua":            '"Google Chrome";v="147", "Not.A/Brand";v="8", "Chromium";v="147"',
    "sec-ch-ua-mobile":     "?0",
    "sec-ch-ua-platform":   '"Windows"',
    "sec-fetch-dest":       "empty",
    "sec-fetch-mode":       "cors",
    "sec-fetch-site":       "same-origin",
    "user-agent":           "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    "x-currency-code":      "INR",
    # x-fp-sdk-version must match what the site expects
    "x-fp-sdk-version":     "1.10.3-60",
}

BB_BASE_HEADERS = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "accept-language": "en-GB,en-US;q=0.9,en;q=0.8",
    "cache-control": "max-age=0", "upgrade-insecure-requests": "1",
    "sec-ch-ua": '"Chromium";v="146", "Google Chrome";v="146"',
    "sec-ch-ua-mobile": "?0", "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "document", "sec-fetch-mode": "navigate", "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
}

DMART_BASE_HEADERS = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-encoding": "gzip, deflate, br, zstd",
    "accept-language": "en-GB,en-US;q=0.9,en;q=0.8",
    "cache-control": "no-cache",
    "pragma": "no-cache",
    "priority": "u=0, i",
    "sec-ch-ua": '"Google Chrome";v="147", "Not.A/Brand";v="8", "Chromium";v="147"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "document",
    "sec-fetch-mode": "navigate",
    "sec-fetch-site": "same-origin",
    "sec-fetch-user": "?1",
    "upgrade-insecure-requests": "1",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/147.0.0.0 Safari/537.36",
    "cookie": '_ga=GA1.1.1223342358.1775106732; uuid=; d_info="w-20260424_120900"; reqId="OTZlMzE2YTQtMTdmNi00MzQ0LWFkYTQtNTYyM2RmMDRmYmZlfHxTLTIwMjYwNDI0XzEyMDkwMHx8LTEwMDI="; guest={"preferredPIN":"520013","preferredStore":"10726","preferredCity":"Vijayawada","preferredArea":"Krishnalanka","isLoggedIn":false,"isPinSet":"true","lat":"16.508267","long":"80.61616169999999","formattedAddress":"Vijayawada bus stand, PN Bus Stand Road, Arrival Block,Pandit Nehru Bus Stop, Krishnalanka, Vijayawada, Andhra Pradesh, India","areaLocationId":""}; recentUserPincodeSearch2=[{"uniqueId":"ChIJQ_rBdgDxNToRzU88N6lTgAU","pincode":"520013","apiMode":"GA","primaryText":"Vijayawada bus stand","secondaryText":"PN Bus Stand Road, Arrival Block,Pandit Nehru Bus Stop, Krishnalanka, Vijayawada, Andhra Pradesh, India","formattedAddress":"Vijayawada bus stand, PN Bus Stand Road, Arrival Block,Pandit Nehru Bus Stop, Krishnalanka, Vijayawada, Andhra Pradesh, India","area":"Krishnalanka","showView":"false","lat":"16.508267","long":"80.61616169999999","newFormattedString":"Vijayawada bus stand, PN Bus Stand Road, Arrival Block,Pandit Nehru Bus Stop, Krishnalanka, Vijayawada, Andhra Pradesh, India","areaLocationId":"","storeId":"10726"}]; _ga_3TR7GSPBGF=GS2.1.s1778136242$o7$g0$t1778136242$j60$l0$h0; _ga_L6YRQ0GLEF=GS2.1.s1778136243$o7$g0$t1778136243$j60$l0$h0'
}

config_store = {
    "spreadsheet_zzz_id": os.getenv("SPREADSHEET_ZZZ_ID","1-GBfTEnnd9_wKRoERB0Z8XURPqh_SU58VG9lMpdAJ5c"),
    "sheet_pmaster":      os.getenv("SHEET_PMASTER", "price master"),
    "credentials_path":   os.getenv("CREDENTIALS_PATH", "credentials.json"),
    "delay_sec":          float(os.getenv("DELAY_SEC", 0.5)),
    "slack_webhook":      os.getenv("SLACK_WEBHOOK", ""),
    "slack_channel":      os.getenv("SLACK_CHANNEL", "#comp-pricing-alerts"),
    # ── JioMart ──────────────────────────────────────────────
    "jio_pincode":        os.getenv("JIO_PINCODE", "516001"),
    "jio_city":           os.getenv("JIO_CITY", "KADAPA"),
    "jio_state":          os.getenv("JIO_STATE", "ANDHRA_PRADESH"),
    "jio_latitude":       os.getenv("JIO_LATITUDE", "14.4783475"),
    "jio_longitude":      os.getenv("JIO_LONGITUDE", "78.821275"),
    "jio_polygon_ids":    os.getenv("JIO_POLYGON_IDS", "6918_QC_72418806,T2ZC_QC_932e287c"),
    "jio_store_ids":      os.getenv("JIO_STORE_IDS", "3121||15594||10319"),
    "jio_authorization":  os.getenv("JIO_AUTHORIZATION", "Bearer Njg1OTQ1ZjQ2YzhjN2FlZTNmM2FmNjA1OlRwS3c3d0Q5aA=="),
    "jio_impersonate":    os.getenv("JIO_IMPERSONATE", "chrome110"),
    # Cookies as JSON — paste the full cookie dict from your working script here,
    # or set individual JIO_COOKIE_* env vars.
    "jio_cookies_json": json.dumps({
        "WZRK_G":               os.getenv("JIO_COOKIE_WZRK_G", "307984d858af48d090f52188245fe5b8"),
        "_fbp":                 os.getenv("JIO_COOKIE_FBP", "fb.1.1763357894292.605657426"),
        "_ALGOLIA":             os.getenv("JIO_COOKIE_ALGOLIA", "anonymous-b8e7e49f-db19-4344-a9ac-77f918dd35ac"),
        "_gcl_au":              os.getenv("JIO_COOKIE_GCL_AU", "1.1.1502610931.1771224018"),
        "_ga":                  os.getenv("JIO_COOKIE_GA", "GA1.1.702012452.1763357895"),
        "nms_mgo_pincode":      os.getenv("JIO_COOKIE_PINCODE", "516001"),
        "nms_mgo_city":         os.getenv("JIO_COOKIE_CITY", "Kadapa"),
        "nms_mgo_state_code":   os.getenv("JIO_COOKIE_STATE_CODE", "AP"),
        "anonymous_id":         os.getenv("JIO_COOKIE_ANON_ID", "ae63f0b303644d198bf9deef4b4ce441"),
        "app_location_details": os.getenv("JIO_COOKIE_LOC_DETAILS",
            '%7B%22country%22%3A%22INDIA%22%2C%22country_iso_code%22%3A%22IN%22%2C%22city%22%3A%22KADAPA%22%2C%22pincode%22%3A%22516001%22%2C%22state%22%3A%22ANDHRA_PRADESH%22%7D'),
        "app_geolocation": os.getenv("JIO_COOKIE_GEOLOCATION",
            '%7B%22latitude%22%3A%2214.4783475%22%2C%22longitude%22%3A%2278.821275%22%2C%22polygon_ids%22%3A%5B%226918_QC_72418806%22%2C%22T2ZC_QC_932e287c%22%5D%7D'),
    }),
    # ── BigBasket ─────────────────────────────────────────────
    "bb_impersonate":     os.getenv("BB_IMPERSONATE", "chrome131"),
    "bb_cookies_json": json.dumps({
    "x-channel":                os.getenv("BB_COOKIE_X_CHANNEL", "web"),
    "_bb_vid":                  os.getenv("BB_COOKIE_BB_VID", ""),
    "_bb_bb2.0":                os.getenv("BB_COOKIE_BB_BB20", "1"),
    "_is_tobacco_enabled":      os.getenv("BB_COOKIE_IS_TOBACCO", "1"),
    "is_subscribe_sa":          os.getenv("BB_COOKIE_IS_SUBSCRIBE_SA", "0"),
    "bigbasket.com":            os.getenv("BB_COOKIE_BIGBASKET", ""),
    "jarvis-id":                os.getenv("BB_COOKIE_JARVIS_ID", ""),
    "ufi":                      os.getenv("BB_COOKIE_UFI", "1"),
    "x-entry-context-id":       os.getenv("BB_COOKIE_X_ENTRY_CTX_ID", "101"),
    "x-entry-context":          os.getenv("BB_COOKIE_X_ENTRY_CTX", "bb-b2b-kirana"),
    "_bb_source":               os.getenv("BB_COOKIE_BB_SOURCE", "pwa"),
    "_client_version":          os.getenv("BB_COOKIE_CLIENT_VERSION", "2843"),
    "_bb_tc":                   os.getenv("BB_COOKIE_BB_TC", "0"),
    "bb2_enabled":              os.getenv("BB_COOKIE_BB2_ENABLED", "true"),
    "xentrycontext":            os.getenv("BB_COOKIE_XENTRY", "bbnow"),
    "xentrycontextid":          os.getenv("BB_COOKIE_XENTRY_ID", "10"),
    "jentrycontextid":          os.getenv("BB_COOKIE_JENTRY_CTX_ID", "10"),
    "isintegratedsa":           os.getenv("BB_COOKIE_IS_INTEGRATED_SA_STR", "true"),
    "is_integrated_sa":         os.getenv("BB_COOKIE_IS_INTEGRATED_SA", "1"),
    "_bb_nhid":                 os.getenv("BB_COOKIE_NHID", ""),
    "_bb_dsid":                 os.getenv("BB_COOKIE_DSID", ""),
    "csrftoken":                os.getenv("BB_COOKIE_CSRFTOKEN", ""),
    "sessionid":                os.getenv("BB_COOKIE_SESSIONID", ""),
    "_bb_hid":                  os.getenv("BB_COOKIE_HID", ""),
    "is_global":                os.getenv("BB_COOKIE_IS_GLOBAL", "0"),
    "_is_bb1.0_supported":      os.getenv("BB_COOKIE_IS_BB10_SUPPORTED", "0"),
    "adb":                      os.getenv("BB_COOKIE_ADB", "0"),
    "csurftoken":               os.getenv("BB_COOKIE_CSURF", ""),
    "_bb_lat_long":             os.getenv("BB_COOKIE_LAT_LONG", ""),
    "_bb_cid":                  os.getenv("BB_COOKIE_CID", ""),
    "_bb_aid":                  os.getenv("BB_COOKIE_AID", ""),
    "_bb_pin_code":             os.getenv("BB_COOKIE_PIN_CODE", "516001"),
    "_bb_sa_ids":               os.getenv("BB_COOKIE_SA_IDS", ""),
    "_bb_cda_sa_info":          os.getenv("BB_COOKIE_CDA_SA_INFO", ""),
    "_ga":                      os.getenv("BB_COOKIE_GA", ""),
}),
    # ── DMart ─────────────────────────────────────────────────
    "dmart_impersonate":  os.getenv("DMART_IMPERSONATE", "chrome131"),
    "dmart_pincode":      os.getenv("DMART_PINCODE", "520013"),
    "dmart_cookies_json": json.dumps({
        "_ga":    os.getenv("DMART_COOKIE_GA", ""),
        "uuid":   os.getenv("DMART_COOKIE_UUID", ""),
        "d_info": os.getenv("DMART_COOKIE_D_INFO", ""),
        "reqId":  os.getenv("DMART_COOKIE_REQ_ID", ""),
    }),
}

# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════

def rlog(msg, level="INFO"):
    run_log.append({"ts": datetime.now().strftime("%H:%M:%S"), "level": level, "msg": str(msg)})

def safe_pct_off(mrp, sp):
    try:
        m, s = float(mrp), float(sp)
        return round((m - s) / m * 100, 2) if m > 0 else None
    except Exception:
        return None

def is_truthy(val):
    return str(val).strip().lower() in ("yes", "y", "true", "1")

ILLEGAL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]")
def _clean(x):
    if x is None: return ""
    s = x if isinstance(x, str) else str(x)
    return re.sub(r"\s+", " ", ILLEGAL_RE.sub("", s)).strip()

def parse_superkids(raw: str) -> list:
    return [p.strip() for p in re.split(r"[,\n;]+", raw or "") if p.strip()]

def _col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def _norm_key(k):
    return re.sub(r"\s+", " ", str(k).strip().lower())


# ══════════════════════════════════════════════════════════════
#  GOOGLE SHEETS
# ══════════════════════════════════════════════════════════════

def get_sheets_service(credentials_path="credentials.json"):
    if not GOOGLE_AVAILABLE:
        raise RuntimeError("google-auth / google-api-python-client not installed.")
    creds = None
    if os.path.exists(TOKEN_PATH):
        try: creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
        except Exception:
            creds = None
            try: os.remove(TOKEN_PATH)
            except OSError: pass
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try: creds.refresh(Request())
            except Exception:
                creds = None
                try: os.remove(TOKEN_PATH)
                except OSError: pass
        if not creds or not creds.valid:
            flow  = InstalledAppFlow.from_client_secrets_file(credentials_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, "w", encoding="utf-8") as f:
            f.write(creds.to_json())
    return build("sheets", "v4", credentials=creds)


def sheet_read_all(service, spreadsheet_id, sheet_name):
    result  = service.spreadsheets().values().get(spreadsheetId=spreadsheet_id, range=sheet_name).execute()
    raw     = result.get("values", [])
    if not raw: return [], [], raw
    headers = [str(h).strip() for h in raw[0]]
    rows    = []
    for row in raw[1:]:
        padded = row + [""] * (len(headers) - len(row))
        rows.append(dict(zip(headers, padded)))
    return headers, rows, raw


def write_range(service, spreadsheet_id, range_a1, values):
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id, range=range_a1,
        valueInputOption="RAW", body={"values": values},
    ).execute()


def ensure_result_columns(service, existing_headers: list) -> dict:
    current = list(existing_headers)
    missing = [c for c in RESULT_COLS if c not in current]
    if missing:
        start = len(current) + 1
        rng   = (f"{MASTER_MAPPER_SHEET}!"
                 f"{_col_letter(start)}1:"
                 f"{_col_letter(start + len(missing) - 1)}1")
        write_range(service, MASTER_MAPPER_ID, rng, [missing])
        current.extend(missing)
        rlog(f"Added result columns: {missing}", "OK")
    return {col: current.index(col) for col in RESULT_COLS if col in current}


def write_result_to_mapper_row(service, row_idx: int, col_map: dict, result: dict):
    if not col_map: return
    min_ci = min(col_map.values())
    max_ci = max(col_map.values())
    cells  = [""] * (max_ci - min_ci + 1)
    for col_name, val in result.items():
        if col_name in col_map:
            cells[col_map[col_name] - min_ci] = "" if val is None else str(val)
    rng = (f"{MASTER_MAPPER_SHEET}!"
           f"{_col_letter(min_ci + 1)}{row_idx}:"
           f"{_col_letter(max_ci + 1)}{row_idx}")
    write_range(service, MASTER_MAPPER_ID, rng, [cells])


def style_diff_cell(service, sheet_id, row_idx, col_0based, diff_val):
    try:
        diff = float(diff_val)
    except (TypeError, ValueError):
        return
    bg = ({"red": 0.988, "green": 0.910, "blue": 0.902} if diff > 0
          else {"red": 0.902, "green": 0.957, "blue": 0.914} if diff < 0
          else None)
    if bg is None: return
    service.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_MAPPER_ID,
        body={"requests": [{"repeatCell": {
            "range": {"sheetId": sheet_id,
                      "startRowIndex": row_idx - 1, "endRowIndex": row_idx,
                      "startColumnIndex": col_0based, "endColumnIndex": col_0based + 1},
            "cell": {"userEnteredFormat": {"backgroundColor": bg}},
            "fields": "userEnteredFormat.backgroundColor",
        }}]},
    ).execute()


def get_mapper_sheet_id(service):
    meta = service.spreadsheets().get(spreadsheetId=MASTER_MAPPER_ID).execute()
    for s in meta["sheets"]:
        if s["properties"]["title"] == MASTER_MAPPER_SHEET:
            return s["properties"]["sheetId"]
    return None


# ══════════════════════════════════════════════════════════════
#  MASTER MAPPER LOADER
# ══════════════════════════════════════════════════════════════

def load_master_mapper(service):
    headers, rows, _ = sheet_read_all(service, MASTER_MAPPER_ID, MASTER_MAPPER_SHEET)
    cleaned = []
    for i, r in enumerate(rows, start=2):
        norm = {k.lower().strip().replace(" ", "_"): str(v).strip() for k, v in r.items()}
        row = {
            "benchmark": norm.get("benchmark", ""),
            "superkid":  norm.get("id","") or norm.get("superkid","") or norm.get("superk_id",""),
            "comp_id":   norm.get("comp_id", ""),
            "url":       norm.get("url", ""),
            "pincode":   norm.get("pincode", ""),
            "_row_idx":  i,
        }
        if not row["superkid"] and not row["comp_id"]:
            continue
        cleaned.append(row)
    return cleaned, headers


# ══════════════════════════════════════════════════════════════
#  PRICE MASTER LOADER
# ══════════════════════════════════════════════════════════════

def load_price_master(service, cfg):
    _, rows, _ = sheet_read_all(service, cfg["spreadsheet_zzz_id"], cfg["sheet_pmaster"])
    default_map = {}
    rlog(f"  Price Master: reading {len(rows)} rows…", "INFO")
    for r in rows:
        norm = {_norm_key(k): str(v).strip() for k, v in r.items()}

        sid = (norm.get("product_id", "") or norm.get("superk_id", "") or
               norm.get("superkid", "") or norm.get("sku_id", ""))
        if not sid: continue

        default_variant_raw = (norm.get("default_variant", "") or norm.get("default variant", ""))
        if not is_truthy(default_variant_raw): continue
        if sid in default_map: continue

        mrp_based_raw = (
                norm.get("mrp_based_pricing", "") or
                norm.get("mrp based pricing", "") or
                norm.get("mrp_based", "")
        ).strip()
        if mrp_based_raw.lower() in ("yes", "y", "true", "1"):
            mrp_based = True
        elif mrp_based_raw.lower() in ("no", "n", "false", "0"):
            mrp_based = False
        else:
            mrp_based = None

        mrp_val = norm.get("mrp", "")
        sp_val = norm.get("selling_price", "") or norm.get("sp", "")
        is_crit = 1 if is_truthy(norm.get("is_critical", "") or norm.get("iscritical", "")) else 0

        # procurement_category — supports multiple column name spellings
        cat = (
                norm.get("procurement_category", "") or
                norm.get("category", "") or
                norm.get("proc_category", "")
        ).strip()

        map_key = sid.lower().strip()

        # ⚠️  FIX: dict created first, category included inside it.
        #    Original had the category assignment on a separate line BEFORE
        #    the dict was initialised → KeyError on every row.
        default_map[map_key] = {
            "mrp": mrp_val,
            "sp": sp_val,
            "mrp_based": mrp_based,
            "is_critical": is_crit,
            "_raw_sid": sid,
            "procurement_category": cat,  # ← now inside the dict literal
        }

    rlog(f"  Price Master: {len(default_map)} superkids loaded", "OK")
    return default_map


# ══════════════════════════════════════════════════════════════
#  SOURCE DETECTION
# ══════════════════════════════════════════════════════════════

def detect_benchmark(benchmark_val, url_val):
    b = str(benchmark_val).strip().lower()
    u = str(url_val).strip().lower()
    if "dmart" in b or "dmart" in u or "dmart.in" in u:
        return "dmart"
    if "jio" in b or "jiomart" in b or "jiomart" in u:
        return "jio"
    if "bb" in b or "bigbasket" in b or "bigbasket" in u:
        return "bb"
    if re.search(r'\bbb\b', b):
        return "bb"
    return "jio"


# ══════════════════════════════════════════════════════════════
#  JIOMART FETCHER  (v4.1 — curl_cffi + cookies, exact headers)
# ══════════════════════════════════════════════════════════════

def _build_jio_headers(cfg, pincode=None):
    """
    Build the full JioMart request headers.

    Key differences from the old version:
    - authorization is used AS-IS from config (the real Bearer token).
    - x-fp-signature is kept as the real value from the working script.
      JioMart doesn't actually validate it server-side when curl_cffi
      impersonation is used, but we include it to look identical.
    - x-fp-date is refreshed per-request (format matches working script).
    - x-geolocation / x-location-detail are built from config values
      exactly as the working script does.
    """
    lat   = cfg.get("jio_latitude",  "14.4783475")
    lng   = cfg.get("jio_longitude", "78.821275")
    # polygon_ids: stored as comma-separated string → convert to list
    raw_pins = cfg.get("jio_polygon_ids", "")
    pins  = [p.strip() for p in raw_pins.split(",") if p.strip()]
    city  = cfg.get("jio_city",  "KADAPA")
    state = cfg.get("jio_state", "ANDHRA_PRADESH")
    pin   = pincode or cfg.get("jio_pincode", "516001")

    hdrs = dict(JIO_BASE_HEADERS)
    hdrs["authorization"]    = cfg.get("jio_authorization", "")
    # Refresh date on every call so it's always current
    hdrs["x-fp-date"]        = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    # Signature: JioMart doesn't revalidate this when impersonation is active;
    # we pass the same value the working script uses so headers look authentic.
    hdrs["x-fp-signature"]   = "v1.1:22724df5123801e07740d09024ac38b0983972874811e86547ee94a1f924de7a"
    hdrs["x-geolocation"]    = json.dumps({
        "latitude":    lat,
        "longitude":   lng,
        "polygon_ids": pins,
    })
    hdrs["x-location-detail"] = json.dumps({
        "country":          "INDIA",
        "country_iso_code": "IN",
        "city":             city,
        "pincode":          pin,
        "state":            state,
    })
    return hdrs


def fetch_jiomart_price(comp_id: str, cfg: dict, pincode: str = None) -> dict:
    """
    Fetch price from JioMart product search API.

    Uses curl_cffi with chrome110 impersonation (matching the working
    standalone script).  Falls back to plain requests if curl_cffi is
    unavailable, but that will likely return empty results.
    """
    if not CURL_AVAILABLE:
        rlog("  ⚠ curl_cffi not installed — JioMart results may be empty", "WARN")

    headers = _build_jio_headers(cfg, pincode)

    # Parse cookies from JSON config
    raw_cookies = cfg.get("jio_cookies_json", "{}")
    try:
        cookies = json.loads(raw_cookies) if isinstance(raw_cookies, str) else raw_cookies
    except Exception:
        cookies = {}
        rlog("  ⚠ Could not parse jio_cookies_json — proceeding without cookies", "WARN")

    # store_ids: kept as pipe-delimited string, e.g. "3121||15594||10319"
    store_ids = cfg.get("jio_store_ids", "")
    params = {
        "f":         f"journey:standard:::store_ids:{store_ids}",
        "page_id":   "*",
        "page_size": 40,   # match working script (was 10 before)
        "q":         comp_id,
    }

    impersonate = cfg.get("jio_impersonate", "chrome110")

    rlog(f"  JIO fetch: comp_id={comp_id} pin={pincode or cfg.get('jio_pincode')} impersonate={impersonate}", "INFO")

    try:
        if CURL_AVAILABLE:
            resp = curl_requests.get(
                JIOMART_SEARCH_URL,
                params=params,
                headers=headers,
                cookies=cookies,
                impersonate=impersonate,
                timeout=30,
            )
        else:
            # Fallback — no impersonation, likely to get blocked
            resp = req_lib.get(
                JIOMART_SEARCH_URL,
                params=params,
                headers=headers,
                cookies=cookies,
                timeout=30,
            )

        resp.raise_for_status()

        try:
            data = resp.json()
        except Exception:
            raise RuntimeError(f"JioMart returned non-JSON (status {resp.status_code}): {resp.text[:200]}")

        # Normalise the response — JioMart has used multiple shapes
        items = (
            data.get("items")
            or data.get("products")
            or data.get("data", {}).get("items")
            or data.get("data", {}).get("products")
            or []
        )

        if not items:
            raise RuntimeError(
                f"JioMart: no products returned for '{comp_id}'. "
                f"Response keys: {list(data.keys())}"
            )

        # Prefer exact UID match, fall back to first result
        item = next((i for i in items if str(i.get("uid", "")) == str(comp_id)), items[0])

        name = item.get("name", "")
        slug = item.get("slug", "")

        price_info     = item.get("price", {})
        effective      = price_info.get("effective", {})
        marked         = price_info.get("marked", {})

        # ── Price extraction (matches working script exactly) ──
        # effective = current / selling price
        # marked    = MRP
        sp_val  = effective.get("min") if effective.get("min") is not None else effective.get("max")
        mrp_val = marked.get("min")    if marked.get("min")    is not None else marked.get("max")

        # Discount percentage
        disc = ""
        if mrp_val and sp_val:
            try:
                disc = str(round(100 - float(sp_val) / float(mrp_val) * 100, 1))
            except Exception:
                pass

        rlog(
            f"  JIO ✓ '{name[:50]}' MRP={mrp_val} SP={sp_val} disc={disc}%",
            "OK" if mrp_val and sp_val else "WARN",
        )

        return {
            "product_name":  name,
            "mrp":           str(mrp_val) if mrp_val is not None else "",
            "selling_price": str(sp_val)  if sp_val  is not None else "",
            "discount_pct":  disc,
            "url":           f"https://www.jiomart.com/p/{slug}" if slug else "",
        }

    except Exception as e:
        rlog(f"  JIO ✗ {e}", "ERROR")
        raise


# ══════════════════════════════════════════════════════════════
#  BIGBASKET FETCHER
# ══════════════════════════════════════════════════════════════

def parse_bb_price_from_html(html):
    result = {"mrp":None,"selling_price":None,"title":"","discount_pct":None}
    try:
        m = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>',html,re.S)
        if m:
            nd   = json.loads(m.group(1))
            pdt  = nd.get("props",{}).get("pageProps",{}).get("productDetails",{})
            prod = (pdt.get("children") or [{}])[0]
            disc = (prod.get("pricing",{}) or {}).get("discount",{}) or {}
            h1   = re.search(r'<h1[^>]*>(.*?)</h1>',html,re.S)
            result["title"] = _clean(h1.group(1)) if h1 else ""
            result["mrp"]   = _clean(disc.get("mrp","")) or None
            sp = ((disc.get("prim_price",{}) or {}).get("sp") or
                  (disc.get("sec_price", {}) or {}).get("sp") or "")
            result["selling_price"] = _clean(sp) or None
            result["discount_pct"]  = disc.get("discount_pct") or disc.get("discount") or None
    except Exception: pass
    if not result["mrp"] or not result["selling_price"]:
        m1=re.search(r'\bMRP\b\D{0,40}₹?\s*([\d,]+(?:\.\d+)?)',html,re.I)
        m2=re.search(r'\b(?:Price|Selling\s*Price|Offer\s*Price)\b\D{0,40}₹?\s*([\d,]+(?:\.\d+)?)',html,re.I)
        if m1 and not result["mrp"]:           result["mrp"]           = m1.group(1)
        if m2 and not result["selling_price"]: result["selling_price"] = m2.group(1)
    return result


def fetch_bb_price(url, cfg):
    try:
        raw_cookies = cfg.get("bb_cookies_json", "{}")
        HEADERS = {
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-language": "en-GB,en-US;q=0.9,en;q=0.8",
            "cache-control": "max-age=0",
            "priority": "u=0, i",
            "sec-ch-ua": '"Chromium";v="146", "Not.A.Brand";v="24", "Google Chrome";v="146"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "same-origin",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
        }

        COOKIES = {
            "x-channel": "web",
            "_bb_vid": "OTkzMTMyODUwMzU1MTkzODM4",
            "_bb_bb2.0": "1",
            "_is_tobacco_enabled": "1",
            "is_subscribe_sa": "0",
            "bigbasket.com": "c15c83ed-9446-4da8-9787-b96e3cca015d",
            "jarvis-id": "82b2d944-3ba8-4f4d-9eac-68accb69217a",
            "ufi": "1",
            "x-entry-context-id": "101",
            "x-entry-context": "bb-b2b-kirana",
            "_bb_source": "pwa",
            "_client_version": "2843",
            "_bb_tc": "0",
            "bb2_enabled": "true",
            "xentrycontext": "bbnow",
            "xentrycontextid": "10",
            "jentrycontextid": "10",
            "isintegratedsa": "true",
            "is_integrated_sa": "1",
            "_bb_nhid": "109",
            "_bb_dsid": "1369",
            "csrftoken": "DPv89d28VTBHpanYBRqNj5ARogKErMXZpI1AyA8a7i2pR1zJSLCi3kqeRk5QbbJo",
            "sessionid": "vmtji0y17krdhiyucbzg511vvq2llvuj",
            "_bb_hid": "7427",
            "is_global": "0",
            "_is_bb1.0_supported": "0",
            "adb": "0",
            "csurftoken": "f9--jA.OTkzMTMyODUwMzU1MTkzODM4.1775123312787.ahbvlRCdHpYI3ZZMDDQrD6CPXcd9mtUKJPiFTbtYlNI=",
            "_bb_lat_long": "MTQuNDc4MzQ3NXw3OC44MjEyNzU=",
            "_bb_cid": "3",
            "_bb_aid": "MzAyMTE5MDE0Ng==",
            "_bb_pin_code": "516001",
            "_bb_sa_ids": "23341",
            "_bb_cda_sa_info": "djIuY2RhX3NhLjEwLjIzMzQx",
            "_ga": "GA1.1.1118534166.1763262544",
        }
        cookies     = json.loads(raw_cookies) if isinstance(raw_cookies, str) else raw_cookies
        impersonate = cfg.get("bb_impersonate", "chrome131")
        if not CURL_AVAILABLE: raise RuntimeError("curl_cffi required for BigBasket")
        if not url: raise RuntimeError("BB requires a url in Master Mapper")
        resp = curl_requests.get(url, headers=HEADERS,
                                  cookies=COOKIES, impersonate="chrome131", timeout=30)
        resp.raise_for_status()
        return parse_bb_price_from_html(resp.text)
    except Exception as e:
        return {"error":str(e),"mrp":None,"selling_price":None,"title":"","discount_pct":None}


# ══════════════════════════════════════════════════════════════
#  DMART FETCHER
# ══════════════════════════════════════════════════════════════

def parse_dmart_price_from_html(html: str) -> dict:
    result = {"mrp": None, "selling_price": None, "title": "", "discount_pct": None}

    # ── Strategy 1: __NEXT_DATA__ ────────────────────────────
    try:
        m = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.S)
        if m:
            nd = json.loads(m.group(1))
            props = nd.get("props", {}).get("pageProps", {})

            def _walk(obj, depth=0):
                if depth > 15: return
                if result["mrp"] and result["selling_price"] and result["title"]: return
                if isinstance(obj, dict):
                    mrp  = obj.get("priceMRP")
                    sp   = obj.get("priceSALE")
                    name = obj.get("name") if (mrp and sp) else None
                    if mrp and sp:
                        result["mrp"]           = str(mrp)
                        result["selling_price"]  = str(sp)
                        if name and not result["title"]:
                            result["title"] = _clean(str(name))
                        return
                    if "sKUs" in obj:
                        _walk(obj["sKUs"], depth + 1)
                    if result["mrp"] and result["selling_price"]: return
                    for k, v in obj.items():
                        if k == "sKUs": continue
                        _walk(v, depth + 1)
                        if result["mrp"] and result["selling_price"]: return
                elif isinstance(obj, list):
                    for v in obj:
                        _walk(v, depth + 1)
                        if result["mrp"] and result["selling_price"]: return

            _walk(props)
    except Exception as e:
        rlog(f"  DMart parse strategy 1 error: {e}", "WARN")

    if result["mrp"] and result["selling_price"]:
        _set_dmart_disc(result)
        return result

    # ── Strategy 2: __next_f push chunks ────────────────────
    try:
        chunks = re.findall(r'self\.__next_f\.push\(\[1\s*,\s*"(.*?)"\]\)', html, re.S)
        combined = ""
        for chunk in chunks:
            try:
                decoded = bytes(chunk, "utf-8").decode("unicode_escape")
            except Exception:
                decoded = chunk
            combined += decoded

        block_re = re.compile(
            r'\{[^{}]{0,3000}?"priceMRP"\s*:\s*"?([\d.]+)"?[^{}]{0,3000}?"priceSALE"\s*:\s*"?([\d.]+)"?[^{}]{0,500}\}',
            re.S
        )
        for blk in block_re.finditer(combined):
            result["mrp"]          = blk.group(1)
            result["selling_price"] = blk.group(2)
            nm = re.search(r'"name"\s*:\s*"([^"]{5,150})"', blk.group(0))
            if nm and not result["title"]:
                result["title"] = _clean(nm.group(1))
            break

        if not (result["mrp"] and result["selling_price"]):
            block_re2 = re.compile(
                r'\{[^{}]{0,3000}?"priceSALE"\s*:\s*"?([\d.]+)"?[^{}]{0,3000}?"priceMRP"\s*:\s*"?([\d.]+)"?[^{}]{0,500}\}',
                re.S
            )
            for blk in block_re2.finditer(combined):
                result["selling_price"] = blk.group(1)
                result["mrp"]          = blk.group(2)
                nm = re.search(r'"name"\s*:\s*"([^"]{5,150})"', blk.group(0))
                if nm and not result["title"]:
                    result["title"] = _clean(nm.group(1))
                break

        if not result["mrp"]:
            mrp_m = re.search(r'"priceMRP"\s*:\s*"?([\d.]+)"?', combined)
            if mrp_m: result["mrp"] = mrp_m.group(1)
        if not result["selling_price"]:
            sale_m = re.search(r'"priceSALE"\s*:\s*"?([\d.]+)"?', combined)
            if sale_m: result["selling_price"] = sale_m.group(1)

    except Exception as e:
        rlog(f"  DMart parse strategy 2 error: {e}", "WARN")

    if result["mrp"] and result["selling_price"]:
        _set_dmart_disc(result)
        return result

    # ── Strategy 3: raw HTML regex fallback ─────────────────
    try:
        mrp_pat = re.search(r'(?:priceMRP|MRP)[^\d]{0,20}([\d,]+(?:\.\d+)?)', html, re.I)
        sp_pat  = re.search(r'(?:priceSALE|Sale Price|Offer Price|Selling Price)[^\d]{0,20}([\d,]+(?:\.\d+)?)', html, re.I)
        h1_pat  = re.search(r'<h1[^>]*>(.*?)</h1>', html, re.S)
        if mrp_pat and not result["mrp"]:          result["mrp"]          = mrp_pat.group(1).replace(",", "")
        if sp_pat  and not result["selling_price"]: result["selling_price"] = sp_pat.group(1).replace(",", "")
        if h1_pat  and not result["title"]:         result["title"]        = _clean(h1_pat.group(1))
    except Exception:
        pass

    _set_dmart_disc(result)
    return result

def _set_dmart_disc(result):
    if result["mrp"] and result["selling_price"]:
        try:
            mrp = float(str(result["mrp"]).replace(",",""))
            sp  = float(str(result["selling_price"]).replace(",",""))
            result["discount_pct"] = round((mrp - sp) / mrp * 100, 1) if mrp > 0 else None
        except Exception:
            pass


def fetch_dmart_price(url: str, cfg: dict) -> dict:
    if not url:
        return {"error": "DMart URL required in Master Mapper (url column)",
                "mrp": None, "selling_price": None, "title": "", "discount_pct": None}
    try:
        raw_cookies = cfg.get("dmart_cookies_json", "{}")
        cookies     = json.loads(raw_cookies) if isinstance(raw_cookies, str) else raw_cookies
        impersonate = cfg.get("dmart_impersonate", "chrome131")

        if CURL_AVAILABLE:
            resp = curl_requests.get(
                url, headers=DMART_BASE_HEADERS, impersonate="chrome120", timeout=30
            )
        elif req_lib:
            resp = req_lib.get(url, headers=DMART_BASE_HEADERS, timeout=30)
        else:
            raise RuntimeError("No HTTP library available (install curl_cffi or requests)")

        resp.raise_for_status()
        parsed = parse_dmart_price_from_html(resp.text)

        if not parsed["mrp"] or not parsed["selling_price"]:
            rlog(f"  DMart: price not found for {url} — HTML snippet: {resp.text[:300]}", "WARN")

        return parsed

    except Exception as e:
        return {"error": str(e), "mrp": None, "selling_price": None, "title": "", "discount_pct": None}


# ══════════════════════════════════════════════════════════════
#  COMPUTE RESULT ROW
# ══════════════════════════════════════════════════════════════

def compute_result(comp_info, sk_entry, run_id, superkid, mrp_based, is_critical):
    ts     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c_mrp  = comp_info.get("mrp", "")
    c_sp   = comp_info.get("selling_price", "")
    status = comp_info.get("status", "OK")
    c_pct  = safe_pct_off(c_mrp, c_sp)

    if mrp_based is True:   mode_label = "MRP"
    elif mrp_based is False: mode_label = "Agnostic"
    else:                   mode_label = ""

    result = {
        "run_id":       run_id,
        "run_at":       ts,
        "product_name": comp_info.get("product_name", ""),
        "is_critical":  is_critical,
        "mode":         mode_label,
        "comp_mrp":     c_mrp,
        "comp_sp":      c_sp,
        "comp_pct_off": "" if c_pct is None else c_pct,
        "sk_mrp":       "",
        "sk_sp":        "",
        "sk_pct_off":   "",
        "diff_pct_off": "",
        "diff_sp":      "",
        "status":       status,
    }

    if not sk_entry:
        return result, f"⚠ '{superkid}' not in Price Master — prices stored, no comparison"

    sk_mrp = sk_entry["mrp"]
    sk_sp  = sk_entry["sp"]
    result["sk_mrp"] = sk_mrp
    result["sk_sp"]  = sk_sp

    if mrp_based is None:
        return result, f"ℹ mode blank for '{superkid}' — prices stored, no diff computed"

    if mrp_based is True:
        sk_pct = safe_pct_off(sk_mrp, sk_sp)
        result["sk_pct_off"] = "" if sk_pct is None else sk_pct
        if status != "OK":
            log_delta = "⚠ skipped diff — fetch error"
        elif c_pct is not None and sk_pct is not None:
            diff = round(c_pct - sk_pct, 2)
            result["diff_pct_off"] = diff
            sign = "RED⚠" if diff > 0 else "GRN✓"
            log_delta = f"[{sign}] comp_pct={c_pct:.2f}% sk_pct={sk_pct:.2f}% diff_pct_off={diff:+.2f}pp"
        elif c_pct is None:
            log_delta = "⚠ MRP: comp % off incalculable"
        else:
            log_delta = "⚠ MRP: SK % off incalculable"
        return result, log_delta
    else:
        if status != "OK":
            log_delta = "⚠ skipped diff — fetch error"
        elif c_sp and sk_sp:
            try:
                diff = round(float(sk_sp) - float(c_sp), 2)
                result["diff_sp"] = diff
                sign = "RED⚠" if diff > 0 else "GRN✓"
                log_delta = f"[{sign}] comp_sp=₹{c_sp} sk_sp=₹{sk_sp} diff_sp=₹{diff:+.2f}"
            except ValueError:
                log_delta = "⚠ Agnostic: non-numeric SP"
        elif not c_sp:
            log_delta = "⚠ Agnostic: comp SP missing"
        else:
            log_delta = "⚠ Agnostic: SK SP missing in Price Master"
        return result, log_delta


# ══════════════════════════════════════════════════════════════
#  CORE RUN LOOP
# ══════════════════════════════════════════════════════════════

def real_run(cfg, filter_superkids=None, source_filter="all"):
    global active_run
    _pause_event.set()
    _stop_event.clear()

    results_map_for_backup = {}  # superkid.lower() → result dict

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode = "selective" if filter_superkids else "full"

    active_run = {
        "running": True, "paused": False, "stopped": False,
        "progress": 0, "total": 0, "current": "",
        "run_id": run_id, "mode": mode,
        "superkids": filter_superkids or [],
        "source_filter": source_filter,
        "last_run_id": run_id,  # NEW: frontend reads this for auto-backup
    }

    start = time.time()
    errors = 0
    written = 0

    rlog("=" * 60)
    rlog(f"PriceLens v5 Run — {run_id}", "INFO")
    rlog(f"Mode: {mode} | Source filter: {source_filter.upper()}", "INFO")
    if filter_superkids:
        rlog(f"SKUs: {filter_superkids}", "INFO")
    rlog("=" * 60)

    try:
        rlog("Authenticating with Google Sheets…", "INFO")
        service = get_sheets_service(cfg["credentials_path"])
        rlog("Auth OK.", "OK")

        rlog("Loading Master Mapper…", "INFO")
        mapper_rows, headers = load_master_mapper(service)
        rlog(f"Mapper: {len(mapper_rows)} rows.", "OK")

        if not mapper_rows:
            rlog("Master Mapper empty — aborting.", "WARN")
            active_run["running"] = False
            return

        if filter_superkids:
            fset = {s.lower() for s in filter_superkids}
            mapper_rows = [r for r in mapper_rows if r["superkid"].lower() in fset]
            rlog(f"Filtered to {len(mapper_rows)} rows for superkids.", "INFO")
            if not mapper_rows:
                rlog("No rows matched the provided Superkids.", "WARN")
                active_run["running"] = False
                return

        if source_filter and source_filter != "all":
            mapper_rows = [
                r for r in mapper_rows
                if detect_benchmark(r["benchmark"], r["url"]) == source_filter.lower()
            ]
            rlog(f"Source filter '{source_filter}': {len(mapper_rows)} rows remaining.", "INFO")
            if not mapper_rows:
                rlog(f"No rows for source '{source_filter}'.", "WARN")
                active_run["running"] = False
                return

        rlog("Ensuring result columns in Master Mapper…", "INFO")
        cur_hdrs, _, _ = sheet_read_all(service, MASTER_MAPPER_ID, MASTER_MAPPER_SHEET)
        col_map = ensure_result_columns(service, cur_hdrs)
        rlog(f"Column map ready ({len(col_map)} cols).", "OK")

        mapper_sheet_id = get_mapper_sheet_id(service)

        rlog("Loading Price Master…", "INFO")
        default_map = load_price_master(service, cfg)
        rlog(f"Price Master: {len(default_map)} superkids loaded.", "OK")

        total = len(mapper_rows)
        active_run["total"] = total
        rlog(f"Processing {total} products…", "INFO")

        for idx, row in enumerate(mapper_rows, start=1):

            if not _pause_event.is_set():
                rlog(f"⏸ PAUSED after {idx - 1} items.", "WARN")
            while not _pause_event.is_set():
                active_run["paused"] = True
                time.sleep(0.3)
            active_run["paused"] = False

            if _stop_event.is_set():
                rlog(f"⛔ STOPPED by user after {idx - 1} items.", "WARN")
                active_run["stopped"] = True
                break

            superkid = row["superkid"]
            comp_id = row["comp_id"]
            url = row["url"]
            pincode = row["pincode"] or cfg.get("jio_pincode", "516001")
            row_idx = row["_row_idx"]
            source = detect_benchmark(row["benchmark"], url)

            active_run["progress"] = idx
            active_run["current"] = f"[{idx}/{total}] {source.upper()} | {superkid}"

            src_icon = {"jio": "🟡", "bb": "🟢", "dmart": "🟠"}.get(source, "⚪")
            rlog(f"[{idx}/{total}] {src_icon} {source.upper()} superkid={superkid} comp_id={comp_id}", "INFO")

            sk_lookup_key = superkid.lower().strip()
            sk_entry = default_map.get(sk_lookup_key)
            if sk_entry:
                mrp_based = sk_entry["mrp_based"]
                is_critical = sk_entry["is_critical"]
                mode_label = "MRP" if mrp_based is True else "Agnostic" if mrp_based is False else "—"
                rlog(f"  ✅ PM: mode={mode_label} sk_mrp={sk_entry['mrp']} sk_sp={sk_entry['sp']}", "INFO")
            else:
                mrp_based = None
                is_critical = 0
                rlog(f"  ❌ '{superkid}' not found in Price Master", "WARN")

            crit_tag = " 🔴CRIT" if is_critical else ""

            comp_info = {}
            try:
                if source == "jio":
                    info = fetch_jiomart_price(comp_id, cfg, pincode)
                    comp_info = {
                        "product_name": info["product_name"],
                        "mrp": info["mrp"],
                        "selling_price": info["selling_price"],
                        "status": "OK"
                    }
                    rlog(f"  JIO '{info['product_name'][:40]}' MRP={info['mrp']} SP={info['selling_price']}", "INFO")

                elif source == "bb":
                    if not url:
                        raise RuntimeError("BB requires url in Master Mapper")
                    info = fetch_bb_price(url, cfg)
                    if "error" in info: raise RuntimeError(info["error"])
                    comp_info = {
                        "product_name": info.get("title", ""),
                        "mrp": str(info.get("mrp") or ""),
                        "selling_price": str(info.get("selling_price") or ""),
                        "status": "OK"
                    }
                    rlog(
                        f"  BB  '{comp_info['product_name'][:40]}' MRP={comp_info['mrp']} SP={comp_info['selling_price']}",
                        "INFO")

                elif source == "dmart":
                    if not url:
                        raise RuntimeError("DMart requires url in Master Mapper")
                    info = fetch_dmart_price(url, cfg)
                    if "error" in info:
                        raise RuntimeError(info["error"])
                    comp_info = {
                        "product_name": info.get("title", ""),
                        "mrp": str(info.get("mrp") or ""),
                        "selling_price": str(info.get("selling_price") or ""),
                        "status": "OK"
                    }
                    if not comp_info["mrp"] or not comp_info["selling_price"]:
                        comp_info["status"] = "ERROR: Price not found in page"
                        errors += 1
                        rlog(f"  ⚠ DMart price extraction failed for {url}", "WARN")
                    else:
                        rlog(
                            f"  DMart '{comp_info['product_name'][:40]}' MRP={comp_info['mrp']} SP={comp_info['selling_price']}",
                            "INFO")
                else:
                    raise RuntimeError(f"Unknown source '{source}'")

            except Exception as e:
                comp_info = {
                    "product_name": "", "mrp": "", "selling_price": "",
                    "status": f"ERROR: {str(e)[:80]}"
                }
                errors += 1
                rlog(f"  ✗ Fetch error: {e}", "ERROR")

            result, log_delta = compute_result(
                comp_info, sk_entry, run_id, superkid, mrp_based, is_critical,
            )

            # ⚠️  FIX: was missing — backup dict was always empty after run
            results_map_for_backup[sk_lookup_key] = result

            if log_delta:
                lvl = "WARN" if "RED" in log_delta else (
                    "ERROR" if ("not found" in log_delta or "incalculable" in log_delta) else "OK"
                )
                rlog(f"  {log_delta}{crit_tag}", lvl)

            try:
                write_result_to_mapper_row(service, row_idx, col_map, result)
                if mapper_sheet_id is not None:
                    diff_key = ("diff_pct_off" if mrp_based is True
                                else "diff_sp" if mrp_based is False else None)
                    diff_val = result.get(diff_key, "") if diff_key else ""
                    if diff_val not in ("", None):
                        col_0b = col_map.get(diff_key, -1)
                        if col_0b >= 0:
                            style_diff_cell(service, mapper_sheet_id, row_idx, col_0b, diff_val)
                written += 1
            except Exception as we:
                rlog(f"  ✗ Sheet write error: {we}", "ERROR")
                errors += 1

            time.sleep(float(cfg.get("delay_sec", 0.5)))

    except Exception as e:
        rlog(f"FATAL: {e}", "ERROR")
        rlog(traceback.format_exc()[-600:], "ERROR")
        errors += 1


    duration = round(time.time() - start, 1)
    ts = datetime.now().strftime("%d %b %Y, %H:%M:%S")
    stopped = _stop_event.is_set()

    rlog("=" * 60)
    rlog(
        f"Run {'STOPPED' if stopped else 'complete'}! "
        f"Written={written} Errors={errors} Duration={duration}s",
        "OK" if not errors else "WARN"
    )

    summary = {
        "run_id": run_id,
        "written": written,
        "errors": errors,
        "duration": f"{duration}s",
        "timestamp": ts,
        "mode": mode,
        "stopped": stopped,
        "superkids": filter_superkids or [],
        "source_filter": source_filter,
    }
    active_run = {
        "running": False,
        "paused": False,
        "stopped": stopped,
        "progress": active_run.get("progress", 0),
        "total": active_run.get("total", 0),
        "run_id": run_id,
        "last_run_id": run_id,  # stays populated after run ends
        "mode": mode,
        "last_result": summary,
        "source_filter": source_filter,
    }
    run_history.insert(0, {**summary, "id": len(run_history) + 1})
    try:
        _send_slack(cfg, summary)
    except Exception:
        pass
    rlog("=" * 60)

# ══════════════════════════════════════════════════════════════
#  SLACK
# ══════════════════════════════════════════════════════════════

def _send_slack(cfg, summary):
    webhook = cfg.get("slack_webhook", "")
    if not webhook or not req_lib:
        return

    sheet_url = f"https://docs.google.com/spreadsheets/d/{MASTER_MAPPER_ID}/edit"
    stopped   = summary.get("stopped", False)
    status_icon = "⛔" if stopped else "✅"
    status_text = "STOPPED" if stopped else "Complete"

    red_count  = 0
    grn_count  = 0
    crit_count = 0
    mrp_diffs  = []
    sp_diffs   = []
    error_count = int(summary.get("errors", 0))

    for entry in list(run_log):
        msg = entry.get("msg", "")
        if "RED⚠" in msg: red_count += 1
        if "GRN✓" in msg: grn_count += 1
        if "🔴CRIT" in msg: crit_count += 1
        m_pct = re.search(r"diff_pct_off=([+-]?[\d.]+)pp", msg)
        if m_pct:
            try: mrp_diffs.append(float(m_pct.group(1)))
            except Exception: pass
        m_sp = re.search(r"diff_sp=₹([+-]?[\d.]+)", msg)
        if m_sp:
            try: sp_diffs.append(float(m_sp.group(1)))
            except Exception: pass

    avg_pct = f"{round(sum(mrp_diffs)/len(mrp_diffs), 2):+.2f}pp" if mrp_diffs else "—"
    avg_sp  = f"₹{round(sum(sp_diffs)/len(sp_diffs), 2):+.2f}"   if sp_diffs  else "—"

    source_upper = (summary.get("source_filter") or "all").upper()
    mode_upper   = (summary.get("mode") or "full").capitalize()
    sks          = summary.get("superkids") or []
    sk_text      = ", ".join(sks[:5]) + ("…" if len(sks) > 5 else "") if sks else "All"

    blocks = [
        {"type": "header", "text": {"type": "plain_text",
            "text": f"{status_icon} PriceLens {status_text} · {summary['run_id']}", "emoji": True}},
        {"type": "section", "fields": [
            {"type": "mrkdwn", "text": f"*Mode*\n{mode_upper}"},
            {"type": "mrkdwn", "text": f"*Source*\n{source_upper}"},
            {"type": "mrkdwn", "text": f"*SKUs*\n{sk_text}"},
            {"type": "mrkdwn", "text": f"*Duration*\n{summary['duration']}"},
            {"type": "mrkdwn", "text": f"*Rows Written*\n{summary['written']}"},
            {"type": "mrkdwn", "text": f"*Errors*\n{'⚠️ ' if error_count else ''}{error_count}"},
        ]},
        {"type": "divider"},
        {"type": "section", "text": {"type": "mrkdwn", "text": (
            "*📈 Price Highlights*\n"
            f"🔴 *Comp cheaper* (comp > SK): `{red_count}`   "
            f"🟢 *SK cheaper* (comp < SK): `{grn_count}`\n"
            f"🔥 *Critical SKUs flagged*: `{crit_count}`\n"
            f"📊 *Avg MRP-mode diff*: `{avg_pct}`   *Avg SP-mode diff*: `{avg_sp}`"
        )}},
        {"type": "divider"},
        {"type": "section", "text": {"type": "mrkdwn",
            "text": f"<{sheet_url}|📋 Open Master Mapper in Google Sheets>"}},
        {"type": "context", "elements": [{"type": "mrkdwn",
            "text": f"PriceLens v4.1 · Run completed at {summary['timestamp']}"}]},
    ]

    try:
        req_lib.post(webhook,
            json={"username": "PriceBot", "icon_emoji": ":bar_chart:", "blocks": blocks},
            timeout=10)
        rlog("Slack sent.", "OK")
    except Exception as e:
        rlog(f"Slack failed: {e}", "WARN")


# ══════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════

_HDR_FILL = PatternFill("solid", fgColor="1A1A2E") if OPENPYXL_AVAILABLE else None
_RED_FILL = PatternFill("solid", fgColor="FFCDD2") if OPENPYXL_AVAILABLE else None
_GRN_FILL = PatternFill("solid", fgColor="C8E6C9") if OPENPYXL_AVAILABLE else None
_ALT_FILL = PatternFill("solid", fgColor="F5F5F5") if OPENPYXL_AVAILABLE else None


def build_excel(headers, data_rows, diff_key, sheet_title="Results"):
    if not OPENPYXL_AVAILABLE: raise RuntimeError("openpyxl not installed")
    dc   = headers.index(diff_key) if diff_key in headers else None
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title         = sheet_title[:31]
    ws.freeze_panes  = "A2"
    white = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _HDR_FILL; c.font = white
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, row in enumerate(data_rows, 2):
        padded = row + [""] * (len(headers) - len(row))
        for ci, val in enumerate(padded, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0: cell.fill = _ALT_FILL
        if dc is not None:
            try:
                dv = float(padded[dc])
                c2 = ws.cell(row=ri, column=dc+1)
                if dv > 0: c2.fill=_RED_FILL; c2.font=Font(name="Calibri",size=9,bold=True,color="C62828")
                elif dv < 0: c2.fill=_GRN_FILL; c2.font=Font(name="Calibri",size=9,bold=True,color="2E7D32")
            except Exception: pass
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 42)
    ws.row_dimensions[1].height = 28
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


# ══════════════════════════════════════════════════════════════
#  FLASK ROUTES
# ══════════════════════════════════════════════════════════════

@app.errorhandler(Exception)
def handle_exception(e):
    return jsonify({"ok":False,"msg":str(e),"trace":traceback.format_exc()[-600:]}), 500

@app.errorhandler(404)
def handle_404(e):
    return jsonify({"ok":False,"msg":f"Not found: {request.path}"}), 404

@app.route("/")
def index():
    return render_template("b.html")


# ── Run control ───────────────────────────────────────────────

@app.route("/api/run", methods=["POST"])
def run_script():
    if active_run.get("running"):
        return jsonify({"ok":False,"msg":"Already running"}), 409
    body          = request.json or {}
    filter_sks    = parse_superkids(body.get("superkids","")) or None
    source_filter = (body.get("source_filter") or "all").lower().strip()
    if source_filter not in ("all","jio","bb","dmart"):
        source_filter = "all"
    threading.Thread(
        target=real_run,
        args=(dict(config_store), filter_sks, source_filter),
        daemon=True
    ).start()
    return jsonify({
        "ok": True,
        "mode": "selective" if filter_sks else "full",
        "superkids": filter_sks or [],
        "source_filter": source_filter,
    })


@app.route("/api/run/pause", methods=["POST"])
def pause_run():
    if not active_run.get("running"):
        return jsonify({"ok":False,"msg":"No run in progress"}), 400
    if _pause_event.is_set():
        _pause_event.clear()
        rlog("⏸ Run PAUSED by user.", "WARN")
        return jsonify({"ok":True,"state":"paused"})
    else:
        _pause_event.set()
        rlog("▶ Run RESUMED by user.", "OK")
        return jsonify({"ok":True,"state":"running"})


@app.route("/api/run/stop", methods=["POST"])
def stop_run():
    if not active_run.get("running"):
        return jsonify({"ok":False,"msg":"No run in progress"}), 400
    _stop_event.set()
    _pause_event.set()
    rlog("⛔ Stop requested by user.", "WARN")
    return jsonify({"ok":True})


@app.route("/api/status")
def status():
    return jsonify({**active_run,
                    "paused": not _pause_event.is_set(),
                    "log_count": len(run_log)})

@app.route("/api/log")
def get_log():
    since = int(request.args.get("since",0))
    return jsonify(list(run_log)[since:])

@app.route("/api/history")
def get_history():
    return jsonify(run_history)


# ── Mapper ────────────────────────────────────────────────────

@app.route("/api/mapper/preview")
def mapper_preview():
    try:
        service = get_sheets_service(config_store["credentials_path"])
        rows, headers = load_master_mapper(service)
        jio   = sum(1 for r in rows if detect_benchmark(r["benchmark"],r["url"])=="jio")
        bb    = sum(1 for r in rows if detect_benchmark(r["benchmark"],r["url"])=="bb")
        dmart = sum(1 for r in rows if detect_benchmark(r["benchmark"],r["url"])=="dmart")
        preview = [{k:v for k,v in r.items() if not k.startswith("_")} for r in rows]
        return jsonify({"ok":True,"total":len(rows),"jio":jio,"bb":bb,"dmart":dmart,
                        "rows":preview,"headers":headers})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})


@app.route("/api/mapper/lookup", methods=["POST"])
def mapper_lookup():
    body    = request.json or {}
    sk_list = parse_superkids(body.get("superkids",""))
    if not sk_list:
        return jsonify({"ok":False,"msg":"No superkids provided"}), 400
    try:
        service = get_sheets_service(config_store["credentials_path"])
        rows, _ = load_master_mapper(service)
        default_map = load_price_master(service, config_store)

        fset    = {s.lower() for s in sk_list}
        matched = [r for r in rows if r["superkid"].lower() in fset]

        results = []
        for r in matched:
            sk  = r["superkid"]
            sk_e = default_map.get(sk, {})
            results.append({
                "superkid":    sk,
                "benchmark":   r["benchmark"],
                "source":      detect_benchmark(r["benchmark"],r["url"]).upper(),
                "comp_id":     r["comp_id"],
                "url":         r["url"],
                "pincode":     r["pincode"],
                "mrp_based":   sk_e.get("mrp_based",False),
                "is_critical": sk_e.get("is_critical",0),
                "sk_mrp":      sk_e.get("mrp",""),
                "sk_sp":       sk_e.get("sp",""),
            })

        not_found = [s for s in sk_list if s.lower() not in {r["superkid"].lower() for r in matched}]
        return jsonify({"ok":True,"found":results,"not_found":not_found,
                        "total_found":len(results)})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})


# ── Results ───────────────────────────────────────────────────

@app.route("/api/results")
def get_results():
    page = int(request.args.get("page", 0))
    limit = int(request.args.get("limit", 50))
    search = (request.args.get("search") or "").lower()
    mode = request.args.get("mode", "all")
    crit = request.args.get("critical", "")
    run_filter = request.args.get("run_id", "")
    source_filter = (request.args.get("source", "") or "").lower()
    cat_filter = (request.args.get("category", "") or "").strip().lower()  # NEW

    try:
        service = get_sheets_service(config_store["credentials_path"])
        _, rows, _ = sheet_read_all(service, MASTER_MAPPER_ID, MASTER_MAPPER_SHEET)
        processed = [r for r in rows if r.get("run_at", "").strip()]

        # mode
        if mode == "mrp":
            processed = [r for r in processed if str(r.get("mode", "")).upper() == "MRP"]
        elif mode == "agnostic":
            processed = [r for r in processed if str(r.get("mode", "")).upper() == "AGNOSTIC"]

        # critical
        if crit == "1":
            processed = [r for r in processed if str(r.get("is_critical", "0")) == "1"]
        elif crit == "0":
            processed = [r for r in processed if str(r.get("is_critical", "0")) == "0"]

        # run_id
        if run_filter:
            processed = [r for r in processed if run_filter in str(r.get("run_id", ""))]

        # source
        if source_filter and source_filter != "all":
            processed = [
                r for r in processed
                if detect_benchmark(
                    r.get("BenchMark", "") or r.get("benchmark", ""),
                    r.get("url", "") or r.get("URL", "")
                ) == source_filter
            ]

        # category (NEW)
        # procurement_category lives in Price Master, not in Master Mapper result rows.
        # We load it once per request and annotate + filter simultaneously.
        default_map = {}
        if cat_filter or True:  # always annotate so Charts tab has the field
            try:
                default_map = load_price_master(service, config_store)
            except Exception as ce:
                rlog(f"get_results: Price Master load error (non-fatal): {ce}", "WARN")

        # annotate every row with its procurement_category
        for r in processed:
            sk_key = (
                    r.get("id", "") or r.get("Superkid", "") or r.get("superkid", "")
            ).lower().strip()
            r["procurement_category"] = default_map.get(sk_key, {}).get("procurement_category", "")

        if cat_filter:
            processed = [
                r for r in processed
                if r.get("procurement_category", "").lower() == cat_filter
            ]

        # text search
        if search:
            processed = [r for r in processed if any(search in str(v).lower() for v in r.values())]

        total = len(processed)
        page_rows = processed[page * limit: (page + 1) * limit]
        return jsonify({"ok": True, "data": page_rows, "total": total})

    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "data": [], "total": 0})


# ── Stats ─────────────────────────────────────────────────────

@app.route("/api/stats")
def stats():
    try:
        service = get_sheets_service(config_store["credentials_path"])
        mapper_rows, _ = load_master_mapper(service)
        jio   = sum(1 for r in mapper_rows if detect_benchmark(r["benchmark"],r["url"])=="jio")
        bb    = sum(1 for r in mapper_rows if detect_benchmark(r["benchmark"],r["url"])=="bb")
        dmart = sum(1 for r in mapper_rows if detect_benchmark(r["benchmark"],r["url"])=="dmart")
        _, rows, _ = sheet_read_all(service, MASTER_MAPPER_ID, MASTER_MAPPER_SHEET)
        processed  = [r for r in rows if r.get("run_at","").strip()]
        mrp_rows   = [r for r in processed if str(r.get("mode","")).upper()=="MRP"]
        ag_rows    = [r for r in processed if str(r.get("mode","")).upper()=="AGNOSTIC"]

        def _agg(rlist, diff_key):
            diffs=[]; crits=0
            for r in rlist:
                v=r.get(diff_key,"")
                if v not in ("",None):
                    try: diffs.append(float(v))
                    except Exception: pass
                if str(r.get("is_critical","0"))=="1": crits+=1
            return {"rows":len(rlist),"red":sum(1 for d in diffs if d>0),
                    "green":sum(1 for d in diffs if d<0),
                    "avg_diff":round(sum(diffs)/len(diffs),2) if diffs else None,
                    "critical":crits}

        return jsonify({"ok":True,
                        "mrp":_agg(mrp_rows,"diff_pct_off"),
                        "agnostic":_agg(ag_rows,"diff_sp"),
                        "mapper":{"total":len(mapper_rows),"jio":jio,"bb":bb,"dmart":dmart}})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})


# ── Critical analysis ─────────────────────────────────────────

@app.route("/api/critical")
def critical_analysis():
    try:
        service = get_sheets_service(config_store["credentials_path"])
        _, rows, _ = sheet_read_all(service, MASTER_MAPPER_ID, MASTER_MAPPER_SHEET)
        processed  = [r for r in rows if r.get("run_at","").strip()]
        crit_rows  = [r for r in processed if str(r.get("is_critical","0"))=="1"]
        mrp_crit   = [r for r in crit_rows  if str(r.get("mode","")).upper()=="MRP"]
        ag_crit    = [r for r in crit_rows  if str(r.get("mode","")).upper()=="AGNOSTIC"]

        def _agg(rlist, diff_key):
            diffs=[]
            for r in rlist:
                try: diffs.append(float(r[diff_key]))
                except Exception: pass
            top10=sorted(rlist,key=lambda x:abs(float(x.get(diff_key) or 0)),reverse=True)[:10]
            return {"count":len(rlist),"red":sum(1 for d in diffs if d>0),
                    "green":sum(1 for d in diffs if d<0),
                    "avg":round(sum(diffs)/len(diffs),2) if diffs else None,
                    "max":max(diffs) if diffs else None,"min":min(diffs) if diffs else None,
                    "top10":top10}

        return jsonify({"ok":True,
                        "mrp":_agg(mrp_crit,"diff_pct_off"),
                        "ag":_agg(ag_crit,"diff_sp"),
                        "mrp_rows":mrp_crit,"ag_rows":ag_crit})
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)})


# ── Excel export ──────────────────────────────────────────────

@app.route("/api/export/excel")
def export_excel():
    mode = request.args.get("mode","all")
    try:
        service    = get_sheets_service(config_store["credentials_path"])
        hdrs, rows,_= sheet_read_all(service, MASTER_MAPPER_ID, MASTER_MAPPER_SHEET)
        processed  = [r for r in rows if r.get("run_at","").strip()]
        if mode=="mrp":      processed=[r for r in processed if str(r.get("mode","")).upper()=="MRP"]
        elif mode=="agnostic":processed=[r for r in processed if str(r.get("mode","")).upper()=="AGNOSTIC"]

        export_cols  = [c for c in hdrs if c in (INPUT_COLS+RESULT_COLS)] or hdrs
        data_rows    = [[str(r.get(c,"")) for c in export_cols] for r in processed]
        diff_key     = "diff_pct_off" if mode=="mrp" else "diff_sp"
        xlsx         = build_excel(export_cols, data_rows, diff_key, f"PriceLens {mode}")
        ts           = datetime.now().strftime("%Y%m%d_%H%M")
        return send_file(io.BytesIO(xlsx),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"pricelens_{mode}_{ts}.xlsx")
    except Exception as e:
        return jsonify({"ok":False,"msg":str(e)}), 500


# ── Config ────────────────────────────────────────────────────

@app.route("/api/config", methods=["GET"])
def get_config():
    return jsonify(config_store)

@app.route("/api/config", methods=["POST"])
def save_config():
    config_store.update(request.json or {})
    rlog("Configuration saved.", "OK")
    return jsonify({"ok":True})


# ── Slack test ────────────────────────────────────────────────

@app.route("/api/test-slack", methods=["POST"])
def test_slack():
    data  = request.json or {}
    dummy = {"run_id":"TEST","written":0,"errors":0,"duration":"test",
             "timestamp":datetime.now().strftime("%d %b %Y, %H:%M:%S"),
             "mode":"full","stopped":False,"superkids":[],"source_filter":"all"}
    _send_slack({**config_store,"slack_webhook":data.get("webhook") or config_store["slack_webhook"]},dummy)
    return jsonify({"ok":True})


# ── Debug price map ───────────────────────────────────────────

@app.route("/api/debug/pricemap")
def debug_pricemap():
    try:
        service  = get_sheets_service(config_store["credentials_path"])
        dm       = load_price_master(service, config_store)
        sk_query = request.args.get("sk", "").strip()
        sk_result = None
        if sk_query:
            sk_result = dm.get(sk_query.lower().strip())
        return jsonify({
            "ok": True, "total_loaded": len(dm),
            "mrp_count": sum(1 for v in dm.values() if v["mrp_based"]),
            "agnostic_count": sum(1 for v in dm.values() if not v["mrp_based"]),
            "query_sk": sk_query or None, "query_result": sk_result,
            "all_keys": list(dm.keys()),
            "sample_entries": {k: v for k, v in list(dm.items())[:5]},
        })
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "trace": traceback.format_exc()[-800:]})


# ── DMart test ────────────────────────────────────────────────

@app.route("/api/test-dmart", methods=["POST"])
def test_dmart():
    body = request.json or {}
    url  = body.get("url","").strip()
    if not url:
        return jsonify({"ok":False,"msg":"URL required"}), 400
    result = fetch_dmart_price(url, config_store)
    return jsonify({"ok": "error" not in result, "url": url, "result": result})

# ── JioMart test  (NEW) ───────────────────────────────────────
@app.route("/api/test-jio", methods=["POST"])
def test_jio():
    """Quick endpoint to test JioMart price fetch for a given comp_id."""
    body    = request.json or {}
    comp_id = body.get("comp_id", "").strip()
    pincode = body.get("pincode", "").strip() or None
    if not comp_id:
        return jsonify({"ok": False, "msg": "comp_id required"}), 400
    try:
        result = fetch_jiomart_price(comp_id, config_store, pincode)
        return jsonify({"ok": True, "comp_id": comp_id, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "comp_id": comp_id, "msg": str(e),
                        "trace": traceback.format_exc()[-600:]})


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════
BACKUP_SHEET_NAME = "PriceLens Backup"

BACKUP_COLS = [
    "backup_run_id", "backed_at", "superkid", "product_name",
    "source", "mode", "procurement_category", "is_critical",
    "comp_mrp", "comp_sp", "comp_pct_off",
    "sk_mrp", "sk_sp", "sk_pct_off",
    "diff_pct_off", "diff_sp", "status", "benchmark",
]


# ══════════════════════════════════════════════════════════════
#  BACKUP SHEET HELPERS
# ══════════════════════════════════════════════════════════════

def ensure_backup_sheet(service):
    """Create the backup sheet if it doesn't exist. Returns sheet_id."""
    meta = service.spreadsheets().get(spreadsheetId=MASTER_MAPPER_ID).execute()
    for s in meta["sheets"]:
        if s["properties"]["title"] == BACKUP_SHEET_NAME:
            return s["properties"]["sheetId"]

    # Create it
    service.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_MAPPER_ID,
        body={"requests": [{"addSheet": {"properties": {"title": BACKUP_SHEET_NAME}}}]},
    ).execute()

    # Write header row
    write_range(service, MASTER_MAPPER_ID,
                f"'{BACKUP_SHEET_NAME}'!A1",
                [BACKUP_COLS])

    # Re-fetch to get sheetId
    meta2 = service.spreadsheets().get(spreadsheetId=MASTER_MAPPER_ID).execute()
    for s in meta2["sheets"]:
        if s["properties"]["title"] == BACKUP_SHEET_NAME:
            return s["properties"]["sheetId"]
    return None


def backup_run_results(service, run_id, mapper_rows, results_map, default_map, cfg):
    """
    Append this run's results to the PriceLens Backup sheet.

    Only SKUs that were actually fetched this run (keys in results_map)
    are written — blank/un-run rows are skipped entirely.

    results_map : dict  superkid.lower() → result dict from compute_result()
    mapper_rows : list  used only to look up benchmark/url for source column
    """
    ensure_backup_sheet(service)
    backed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Build a fast lookup: superkid.lower() → mapper row
    mapper_lookup = {
        r.get("superkid", "").lower().strip(): r
        for r in mapper_rows
        if r.get("superkid", "").strip()
    }

    rows_to_append = []

    # ── iterate only over SKUs that were actually run ──────────
    for sk_key, result in results_map.items():
        if not sk_key:
            continue

        mapper_row = mapper_lookup.get(sk_key, {})
        sk_entry = default_map.get(sk_key, {})

        source = detect_benchmark(
            mapper_row.get("benchmark", ""),
            mapper_row.get("url", ""),
        ).upper()

        cat = sk_entry.get("procurement_category", "")

        row = [
            run_id,
            backed_at,
            result.get("_raw_sk") or sk_key,  # original-case superkid if stored, else key
            result.get("product_name", ""),
            source,
            result.get("mode", ""),
            cat,
            result.get("is_critical", 0),
            result.get("comp_mrp", ""),
            result.get("comp_sp", ""),
            result.get("comp_pct_off", ""),
            result.get("sk_mrp", ""),
            result.get("sk_sp", ""),
            result.get("sk_pct_off", ""),
            result.get("diff_pct_off", ""),
            result.get("diff_sp", ""),
            result.get("status", ""),
            mapper_row.get("benchmark", ""),
        ]
        rows_to_append.append([str(v) if v is not None else "" for v in row])

    if not rows_to_append:
        rlog("Backup: nothing to write (no SKUs were run this session).", "INFO")
        return 0

    service.spreadsheets().values().append(
        spreadsheetId=MASTER_MAPPER_ID,
        range=f"'{BACKUP_SHEET_NAME}'!A1",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": rows_to_append},
    ).execute()

    rlog(f"Backup: {len(rows_to_append)} rows written (ran SKUs only).", "OK")
    return len(rows_to_append)


# ══════════════════════════════════════════════════════════════
#  LOAD CATEGORIES from Price Master
# ══════════════════════════════════════════════════════════════

def load_categories(service, cfg):
    """Return sorted unique procurement_category values from Price Master."""
    _, rows, _ = sheet_read_all(service, cfg["spreadsheet_zzz_id"], cfg["sheet_pmaster"])
    cats = set()
    for r in rows:
        norm = {k.lower().replace(" ", "_"): str(v).strip() for k, v in r.items()}
        cat = (
                norm.get("procurement_category", "")
                or norm.get("category", "")
                or norm.get("proc_category", "")
        ).strip()
        if cat:
            cats.add(cat)
    return sorted(cats)


# ══════════════════════════════════════════════════════════════
#  LOAD TRENDS from Backup Sheet
# ══════════════════════════════════════════════════════════════

def load_trends_data(service, filters=None):
    """
    Load backup sheet and return time series per SKU.
    filters: dict with optional keys: superkids[], category, source, mode, days
    """
    filters = filters or {}
    try:
        result = service.spreadsheets().values().get(
            spreadsheetId=MASTER_MAPPER_ID,
            range=BACKUP_SHEET_NAME,
        ).execute()
        raw = result.get("values", [])
    except Exception as e:
        return {"error": str(e), "runs": [], "skus": {}}

    if not raw or len(raw) < 2:
        return {"runs": [], "skus": {}, "categories": [], "total_rows": 0}

    headers = [str(h).strip() for h in raw[0]]
    rows = []
    for row in raw[1:]:
        padded = row + [""] * (len(headers) - len(row))
        rows.append(dict(zip(headers, padded)))

    # Apply filters
    sk_filter = {s.lower() for s in (filters.get("superkids") or [])}
    cat_filter = (filters.get("category") or "").lower()
    src_filter = (filters.get("source") or "").lower()
    mode_filter = (filters.get("mode") or "").upper()
    days = int(filters.get("days") or 90)

    from datetime import timedelta
    cutoff = datetime.now() - timedelta(days=days)

    filtered = []
    for r in rows:
        backed_at = r.get("backed_at", "")
        try:
            row_dt = datetime.strptime(backed_at[:19], "%Y-%m-%d %H:%M:%S")
            if row_dt < cutoff:
                continue
        except Exception:
            pass

        if sk_filter and r.get("superkid", "").lower() not in sk_filter:
            continue
        if cat_filter and cat_filter != "all":
            rc = r.get("procurement_category", "").lower()
            if rc != cat_filter:
                continue
        if src_filter and src_filter != "all":
            rs = r.get("source", "").lower()
            if src_filter not in rs:
                continue
        if mode_filter and mode_filter not in ("ALL", ""):
            rm = r.get("mode", "").upper()
            if rm != mode_filter:
                continue
        filtered.append(r)

    # Build per-SKU time series
    sku_map = {}
    for r in filtered:
        sk = r.get("superkid", "")
        if not sk:
            continue
        if sk not in sku_map:
            sku_map[sk] = {
                "superkid": sk,
                "product_name": r.get("product_name", ""),
                "category": r.get("procurement_category", ""),
                "source": r.get("source", ""),
                "mode": r.get("mode", ""),
                "is_critical": r.get("is_critical", "0"),
                "points": [],
            }
        try:
            diff_pct = float(r.get("diff_pct_off", "") or "nan")
        except Exception:
            diff_pct = None
        try:
            diff_sp = float(r.get("diff_sp", "") or "nan")
        except Exception:
            diff_sp = None
        try:
            comp_sp = float(r.get("comp_sp", "") or "nan")
        except Exception:
            comp_sp = None
        try:
            sk_sp = float(r.get("sk_sp", "") or "nan")
        except Exception:
            sk_sp = None
        try:
            comp_mrp = float(r.get("comp_mrp", "") or "nan")
        except Exception:
            comp_mrp = None

        import math
        sku_map[sk]["points"].append({
            "run_id": r.get("backup_run_id", ""),
            "date": r.get("backed_at", ""),
            "diff_pct": None if diff_pct is None or (
                        isinstance(diff_pct, float) and math.isnan(diff_pct)) else diff_pct,
            "diff_sp": None if diff_sp is None or (isinstance(diff_sp, float) and math.isnan(diff_sp)) else diff_sp,
            "comp_sp": None if comp_sp is None or (isinstance(comp_sp, float) and math.isnan(comp_sp)) else comp_sp,
            "sk_sp": None if sk_sp is None or (isinstance(sk_sp, float) and math.isnan(sk_sp)) else sk_sp,
            "comp_mrp": None if comp_mrp is None or (
                        isinstance(comp_mrp, float) and math.isnan(comp_mrp)) else comp_mrp,
            "status": r.get("status", ""),
        })

    # Sort points by date
    for sk_data in sku_map.values():
        sk_data["points"].sort(key=lambda p: p["date"])

    # Unique run_ids in order
    run_ids = sorted({r.get("backup_run_id", "") for r in filtered if r.get("backup_run_id")})
    cats = sorted({r.get("procurement_category", "") for r in filtered if r.get("procurement_category")})

    return {
        "runs": run_ids,
        "skus": sku_map,
        "categories": cats,
        "total_rows": len(filtered),
    }


# ══════════════════════════════════════════════════════════════
#  FLASK ROUTES — add these to app.py
# ══════════════════════════════════════════════════════════════

@app.route("/api/categories")
def get_categories():
    try:
        service = get_sheets_service(config_store["credentials_path"])
        cats = load_categories(service, config_store)
        return jsonify({"ok": True, "categories": cats})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "categories": []})


@app.route("/api/backup/run", methods=["POST"])
def manual_backup():
    """Manually trigger a backup for a specific run_id only."""
    try:
        service = get_sheets_service(config_store["credentials_path"])
        mapper_rows, _ = load_master_mapper(service)
        default_map = load_price_master(service, config_store)
        _, result_rows, _ = sheet_read_all(service, MASTER_MAPPER_ID, MASTER_MAPPER_SHEET)

        run_id = (request.json or {}).get("run_id") or datetime.now().strftime("MANUAL_%Y%m%d_%H%M%S")

        # Only rows that belong to this specific run_id
        results_map = {}
        for r in result_rows:
            if r.get("run_id", "").strip() != run_id:
                continue
            sk = (r.get("Superkid") or r.get("superkid") or r.get("id") or "").lower().strip()
            if sk:
                results_map[sk] = {
                    "product_name": r.get("product_name", ""),
                    "mode": r.get("mode", ""),
                    "is_critical": r.get("is_critical", "0"),
                    "comp_mrp": r.get("comp_mrp", ""),
                    "comp_sp": r.get("comp_sp", ""),
                    "comp_pct_off": r.get("comp_pct_off", ""),
                    "sk_mrp": r.get("sk_mrp", ""),
                    "sk_sp": r.get("sk_sp", ""),
                    "sk_pct_off": r.get("sk_pct_off", ""),
                    "diff_pct_off": r.get("diff_pct_off", ""),
                    "diff_sp": r.get("diff_sp", ""),
                    "status": r.get("status", ""),
                }

        if not results_map:
            return jsonify({"ok": False, "msg": f"No rows found for run_id '{run_id}'"})

        n = backup_run_results(service, run_id, mapper_rows, results_map, default_map, config_store)
        rlog(f"Backup: {n} rows written to '{BACKUP_SHEET_NAME}'", "OK")
        return jsonify({"ok": True, "rows_written": n, "run_id": run_id})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "trace": traceback.format_exc()[-600:]})

@app.route("/api/trends")
def get_trends():
    """Get historical time series from backup sheet."""
    filters = {
        "superkids": request.args.getlist("sk") or [],
        "category": request.args.get("category", ""),
        "source": request.args.get("source", ""),
        "mode": request.args.get("mode", ""),
        "days": request.args.get("days", 90),
    }
    try:
        service = get_sheets_service(config_store["credentials_path"])
        data = load_trends_data(service, filters)
        return jsonify({"ok": True, **data})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "runs": [], "skus": {}, "categories": []})


@app.route("/api/trends/runs")
def get_trend_runs():
    """List all unique run_ids in backup sheet with counts."""
    try:
        service = get_sheets_service(config_store["credentials_path"])
        result = service.spreadsheets().values().get(
            spreadsheetId=MASTER_MAPPER_ID, range=BACKUP_SHEET_NAME,
        ).execute()
        raw = result.get("values", [])
        if not raw or len(raw) < 2:
            return jsonify({"ok": True, "runs": []})
        headers = [str(h).strip() for h in raw[0]]
        run_col = headers.index("backup_run_id") if "backup_run_id" in headers else 0
        date_col = headers.index("backed_at") if "backed_at" in headers else 1
        runs = {}
        for row in raw[1:]:
            if len(row) <= run_col: continue
            rid = row[run_col]
            dat = row[date_col] if len(row) > date_col else ""
            if rid not in runs:
                runs[rid] = {"run_id": rid, "date": dat[:16], "count": 0}
            runs[rid]["count"] += 1
        sorted_runs = sorted(runs.values(), key=lambda x: x["date"], reverse=True)
        return jsonify({"ok": True, "runs": sorted_runs})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "runs": []})
if __name__ == "__main__":
    print("=" * 60)
    print("PriceLens v4.1 (JioMart + BigBasket + DMart) → http://localhost:5050")
    print(f"  Master Mapper ID   : {MASTER_MAPPER_ID}")
    print(f"  Master Mapper Sheet: {MASTER_MAPPER_SHEET}")
    print("  Sources: JioMart 🟡 · BigBasket 🟢 · DMart Ready 🟠")
    print("  JioMart: curl_cffi chrome110 impersonation + cookies")
    print("  Features: Pause/Stop · Selective · Source Filter · Lookup")
    print("=" * 60)
    app.run(host='10.20.8.177',debug=False, port=5050, threaded=True)